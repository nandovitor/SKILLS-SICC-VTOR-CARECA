#!/usr/bin/env python3
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path


def emit(payload):
    sys.stdout.write(json.dumps(payload, ensure_ascii=False))


def read_pdf(file_path):
    from pypdf import PdfReader

    reader = PdfReader(file_path)
    chunks = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text:
            chunks.append(text)
    return "\n".join(chunks), "python-pypdf"


def read_docx(file_path):
    from docx import Document

    document = Document(file_path)
    lines = [paragraph.text for paragraph in document.paragraphs if paragraph.text]
    return "\n".join(lines), "python-docx"


def read_xlsx(file_path):
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True)
    lines = []

    for worksheet in workbook.worksheets:
        lines.append(f"[sheet] {worksheet.title}")
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if values:
                lines.append(" | ".join(values))

    return "\n".join(lines), "python-openpyxl"


def read_xls(file_path):
    import xlrd

    workbook = xlrd.open_workbook(file_path)
    lines = []

    for sheet in workbook.sheets():
        lines.append(f"[sheet] {sheet.name}")
        for row_index in range(sheet.nrows):
            row = sheet.row_values(row_index)
            values = [str(value).strip() for value in row if str(value).strip()]
            if values:
                lines.append(" | ".join(values))

    return "\n".join(lines), "python-xlrd"


def read_doc_with_textract(file_path):
    import textract

    raw = textract.process(file_path)
    return raw.decode("utf-8", errors="ignore"), "python-textract"


def read_doc_with_word(file_path):
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False
    document = None

    try:
        document = word.Documents.Open(str(file_path), ReadOnly=True)
        text = document.Content.Text or ""
        return text, "python-win32com"
    finally:
        if document is not None:
            document.Close(False)
        word.Quit()
        pythoncom.CoUninitialize()


def read_doc_with_antiword(file_path):
    antiword_path = shutil.which("antiword")
    if not antiword_path:
        raise RuntimeError("antiword nao encontrado no PATH")

    result = subprocess.run(
        [antiword_path, str(file_path)],
        capture_output=True,
        text=True,
        check=True,
    )
    return result.stdout, "antiword"


def extract(file_path):
    suffix = Path(file_path).suffix.lower()

    if suffix == ".pdf":
        return read_pdf(file_path)
    if suffix == ".docx":
        return read_docx(file_path)
    if suffix == ".xlsx":
        return read_xlsx(file_path)
    if suffix == ".xls":
        return read_xls(file_path)
    if suffix == ".doc":
        errors = []

        try:
            return read_doc_with_textract(file_path)
        except Exception as exc:
            errors.append(f"textract: {exc}")

        if sys.platform.startswith("win"):
            try:
                return read_doc_with_word(file_path)
            except Exception as exc:
                errors.append(f"win32com: {exc}")

        try:
            return read_doc_with_antiword(file_path)
        except Exception as exc:
            errors.append(f"antiword: {exc}")

        raise RuntimeError(" ; ".join(errors))

    raise RuntimeError(f"Formato nao suportado pelo helper Python: {suffix or 'sem extensao'}")


def main():
    if len(sys.argv) < 2:
        emit({"ok": False, "error": "Arquivo nao informado."})
        return 1

    file_path = os.path.abspath(sys.argv[1])

    try:
        text, engine = extract(file_path)
        emit({
            "ok": True,
            "engine": engine,
            "text": (text or "").replace("\r\n", "\n").strip(),
        })
        return 0
    except Exception as exc:
        emit({
            "ok": False,
            "error": str(exc),
        })
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
